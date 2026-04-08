# backend/apps/appointments/views.py
import io
from datetime import timedelta

from django.db import connection, transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.authentication.permissions import IsStaffOrAdmin
from apps.customers.auth import CustomerJWTAuthentication
from apps.customers.permissions import IsAuthenticatedCustomer
from apps.vehicles.models import Vehicle

from .models import Appointment, AppointmentSlot
from .serializers import (
    AppointmentSerializer,
    AppointmentSlotCustomerSerializer,
    AppointmentSlotSerializer,
)

CLOSED_FOR_CAPACITY = ("cancelled",)
ALLOWED_STATUSES = ("scheduled", "confirmed", "in_progress", "completed", "cancelled")


def _count_used_capacity(slot_id: str) -> int:
    with connection.cursor() as cursor:
        cursor.execute(
            """
            select count(*)
            from public.appointments
            where slot_id = %s
              and coalesce(status,'') not in ('cancelled')
            """,
            [slot_id],
        )
        return int(cursor.fetchone()[0] or 0)


class AppointmentSlotAdminViewSet(viewsets.ModelViewSet):
    serializer_class = AppointmentSlotSerializer
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]

    def get_queryset(self):
        qs = AppointmentSlot.objects.all().order_by("start_at")
        is_active = self.request.query_params.get("is_active")
        if is_active is not None and str(is_active).lower() in ("true", "false"):
            qs = qs.filter(is_active=(str(is_active).lower() == "true"))
        return qs

    def create(self, request, *args, **kwargs):
        data = request.data or {}
        start_at = data.get("start_at")
        end_at = data.get("end_at") or None
        capacity = data.get("capacity")
        is_active = data.get("is_active", True)
        notes = (data.get("notes") or "").strip() or None

        if not start_at:
            return Response({"detail": "start_at es requerido."}, status=400)
        if capacity is None:
            return Response({"detail": "capacity es requerido."}, status=400)

        try:
            capacity_int = int(capacity)
            if capacity_int < 1:
                return Response({"detail": "capacity debe ser >= 1."}, status=400)
        except Exception:
            return Response({"detail": "capacity inválido."}, status=400)

        with connection.cursor() as cursor:
            cursor.execute(
                """
                insert into public.appointment_slots
                  (start_at, end_at, capacity, is_active, notes, created_at, updated_at)
                values
                  (%s, %s, %s, %s, %s, now(), now())
                returning slot_id
                """,
                [start_at, end_at, capacity_int, bool(is_active), notes],
            )
            slot_id = cursor.fetchone()[0]

        slot = AppointmentSlot.objects.get(slot_id=slot_id)
        return Response(self.get_serializer(slot).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        slot = self.get_object()
        data = request.data or {}

        sets = []
        params = []

        if "start_at" in data:
            v = data.get("start_at")
            if not v:
                return Response({"detail": "start_at no puede ser vacío."}, status=400)
            sets.append("start_at = %s")
            params.append(v)

        if "end_at" in data:
            sets.append("end_at = %s")
            params.append(data.get("end_at") or None)

        if "capacity" in data:
            try:
                cap = int(data.get("capacity"))
                if cap < 1:
                    return Response({"detail": "capacity debe ser >= 1."}, status=400)
            except Exception:
                return Response({"detail": "capacity inválido."}, status=400)
            sets.append("capacity = %s")
            params.append(cap)

        if "is_active" in data:
            sets.append("is_active = %s")
            params.append(bool(data.get("is_active")))

        if "notes" in data:
            sets.append("notes = %s")
            params.append((data.get("notes") or "").strip() or None)

        if not sets:
            slot.refresh_from_db()
            return Response(self.get_serializer(slot).data, status=200)

        sets.append("updated_at = now()")
        params.append(str(slot.slot_id))

        with connection.cursor() as cursor:
            cursor.execute(
                f"update public.appointment_slots set {', '.join(sets)} where slot_id = %s",
                params,
            )

        slot.refresh_from_db()
        return Response(self.get_serializer(slot).data, status=200)

    def destroy(self, request, *args, **kwargs):
        slot = self.get_object()
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT COUNT(*) FROM public.appointments WHERE slot_id = %s AND coalesce(status,'') != 'cancelled'",
                [str(slot.slot_id)],
            )
            active_count = cursor.fetchone()[0]
        if active_count:
            return Response(
                {"detail": f"No se puede eliminar: el slot tiene {active_count} cita(s) activa(s)."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        with connection.cursor() as cursor:
            cursor.execute("delete from public.appointment_slots where slot_id = %s", [str(slot.slot_id)])
        return Response(status=status.HTTP_204_NO_CONTENT)


class CustomerSlotViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = AppointmentSlotCustomerSerializer
    authentication_classes = [CustomerJWTAuthentication]
    permission_classes = [IsAuthenticatedCustomer]

    def get_queryset(self):
        now = timezone.now()
        return AppointmentSlot.objects.filter(is_active=True, start_at__gte=now).order_by("start_at")

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        rows = list(qs)

        result = []
        for s in rows:
            used = _count_used_capacity(str(s.slot_id))
            remaining = max(int(s.capacity) - used, 0)
            if remaining <= 0:
                continue
            result.append(
                {
                    "slot_id": str(s.slot_id),
                    "start_at": s.start_at,
                    "end_at": s.end_at,
                    "remaining_capacity": remaining,
                }
            )
        return Response(result, status=200)


class AppointmentAdminViewSet(viewsets.ModelViewSet):
    serializer_class = AppointmentSerializer
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]

    def get_queryset(self):
        qs = Appointment.objects.select_related("customer", "vehicle", "service", "slot").prefetch_related("work_orders").all().order_by("-scheduled_start")

        status_q = self.request.query_params.get("status")
        if status_q:
            qs = qs.filter(status__iexact=str(status_q).strip())

        customer_id = self.request.query_params.get("customer_id")
        if customer_id:
            qs = qs.filter(customer_id=customer_id)

        vehicle_id = self.request.query_params.get("vehicle_id")
        if vehicle_id:
            qs = qs.filter(vehicle_id=vehicle_id)

        service_id = self.request.query_params.get("service_id")
        if service_id:
            qs = qs.filter(service_id=service_id)

        slot_id = self.request.query_params.get("slot_id")
        if slot_id:
            qs = qs.filter(slot_id=slot_id)

        return qs

    def partial_update(self, request, *args, **kwargs):
        ap = self.get_object()
        data = request.data or {}

        sets = []
        params = []

        if "status" in data:
            v = (data.get("status") or "").strip()
            if v not in ALLOWED_STATUSES:
                return Response({"detail": "status inválido."}, status=400)
            sets.append("status = %s")
            params.append(v)

        if "admin_message" in data:
            sets.append("admin_message = %s")
            params.append((data.get("admin_message") or "").strip() or None)

        if "progress_percent" in data:
            try:
                p = int(data.get("progress_percent"))
            except Exception:
                return Response({"detail": "progress_percent inválido."}, status=400)
            if p < 0 or p > 100:
                return Response({"detail": "progress_percent debe estar entre 0 y 100."}, status=400)
            sets.append("progress_percent = %s")
            params.append(p)

        if "scheduled_end" in data:
            sets.append("scheduled_end = %s")
            params.append(data.get("scheduled_end") or None)

        if "assigned_mechanic_id" in data:
            sets.append("assigned_mechanic_id = %s")
            params.append(data.get("assigned_mechanic_id") or None)

        if not sets:
            ap = Appointment.objects.select_related("customer", "vehicle", "service", "slot").get(appointment_id=ap.appointment_id)
            return Response(self.get_serializer(ap).data, status=200)

        sets.append("updated_at = now()")
        params.append(str(ap.appointment_id))

        with connection.cursor() as cursor:
            cursor.execute(
                f"update public.appointments set {', '.join(sets)} where appointment_id = %s",
                params,
            )

        ap.refresh_from_db()
        ap = Appointment.objects.select_related("customer", "vehicle", "service", "slot").get(appointment_id=ap.appointment_id)
        return Response(self.get_serializer(ap).data, status=200)

    @transaction.atomic
    @action(detail=True, methods=["post"], url_path="confirm")
    def confirm(self, request, pk=None):
        """Confirma la cita y crea automáticamente una OT vinculada si no existe.
        Funciona desde cualquier estado activo (scheduled, confirmed, in_progress)."""
        ap = self.get_object()

        if ap.status == "cancelled":
            return Response(
                {"detail": "No se puede operar sobre una cita cancelada."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        work_order_id = None
        with connection.cursor() as cursor:
            # Verificar si ya existe una OT vinculada
            cursor.execute(
                "SELECT work_order_id FROM public.work_orders WHERE appointment_id = %s LIMIT 1",
                [str(ap.appointment_id)],
            )
            row = cursor.fetchone()
            if row:
                work_order_id = str(row[0])
            else:
                # Crear OT con SQL puro (consistente con el resto del módulo)
                symptoms = (ap.requested_work or "").strip() or None
                mechanic = str(ap.assigned_mechanic_id) if ap.assigned_mechanic_id else None
                created_by = str(request.user.id)
                cursor.execute(
                    """
                    INSERT INTO public.work_orders
                      (appointment_id, customer_id, vehicle_id, status,
                       customer_symptoms, assigned_mechanic_id, created_by,
                       opened_at, created_at, updated_at)
                    VALUES (%s, %s, %s, 'open', %s, %s, %s, now(), now(), now())
                    RETURNING work_order_id
                    """,
                    [
                        str(ap.appointment_id),
                        str(ap.customer_id),
                        str(ap.vehicle_id),
                        symptoms,
                        mechanic,
                        created_by,
                    ],
                )
                work_order_id = str(cursor.fetchone()[0])

            # Actualizar estado de la cita solo si está en "scheduled"
            if ap.status == "scheduled":
                cursor.execute(
                    "UPDATE public.appointments SET status = 'confirmed', updated_at = now() WHERE appointment_id = %s",
                    [str(ap.appointment_id)],
                )

        ap = Appointment.objects.select_related("customer", "vehicle", "service", "slot").get(
            appointment_id=ap.appointment_id
        )
        return Response(
            {
                "appointment": self.get_serializer(ap).data,
                "work_order_id": work_order_id,
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=False, methods=["get"], url_path="report")
    def report(self, request):
        """Reporte de citas con filtros por fecha, estado y servicio.
        Soporta format=excel para exportar a .xlsx (openpyxl)."""
        params = request.query_params
        now = timezone.now()

        # Defaults: semana actual
        today = now.date()
        week_start = today - timedelta(days=today.weekday())
        date_from = params.get("date_from") or str(week_start)
        date_to = params.get("date_to") or str(today)
        status_filter = (params.get("status") or "").strip()
        service_id_filter = (params.get("service_id") or "").strip()
        export_format = (params.get("export") or "json").strip().lower()

        filters = ["a.scheduled_start::date >= %s", "a.scheduled_start::date <= %s"]
        values = [date_from, date_to]

        if status_filter:
            filters.append("a.status = %s")
            values.append(status_filter)

        if service_id_filter:
            filters.append("a.service_id::text = %s")
            values.append(service_id_filter)

        where_clause = " and ".join(filters)

        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                select
                    a.appointment_id,
                    a.scheduled_start,
                    a.scheduled_end,
                    a.status,
                    a.requested_work,
                    a.notes,
                    a.admin_message,
                    c.full_name as customer_name,
                    c.email as customer_email,
                    v.plate as vehicle_plate,
                    v.make as vehicle_make,
                    v.model as vehicle_model,
                    s.name as service_name,
                    u.username as mechanic_username
                from public.appointments a
                left join public.customers c on c.customer_id = a.customer_id
                left join public.vehicles v on v.vehicle_id = a.vehicle_id
                left join public.services s on s.service_id = a.service_id
                left join django_app.auth_users u on u.id = a.assigned_mechanic_id
                where {where_clause}
                order by a.scheduled_start
                """,
                values,
            )
            rows = cursor.fetchall()
            cols = [d[0] for d in cursor.description]

        appointments = [dict(zip(cols, row)) for row in rows]

        # Serializar campos no-JSON
        for ap in appointments:
            for k, v in ap.items():
                if hasattr(v, "isoformat"):
                    ap[k] = v.isoformat()
                elif v is None:
                    ap[k] = None
                else:
                    ap[k] = str(v) if not isinstance(v, (str, int, float, bool)) else v

        # Resumen por estado
        by_status = {}
        for ap in appointments:
            st = ap.get("status") or "unknown"
            by_status[st] = by_status.get(st, 0) + 1

        # Resumen por día (dentro del rango)
        day_names = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
        day_counts = {}
        for ap in appointments:
            raw = ap.get("scheduled_start") or ""
            if raw:
                try:
                    from datetime import datetime
                    dt = datetime.fromisoformat(raw)
                    d = dt.date()
                    key = str(d)
                    if key not in day_counts:
                        day_counts[key] = {"date": key, "day": day_names[d.weekday()], "count": 0}
                    day_counts[key]["count"] += 1
                except Exception:
                    pass

        by_day = sorted(day_counts.values(), key=lambda x: x["date"])

        summary = {
            "total": len(appointments),
            "by_status": by_status,
            "by_day": by_day,
        }

        if export_format == "excel":
            return _build_excel_response(appointments, by_day, date_from, date_to)

        return Response({"appointments": appointments, "summary": summary}, status=200)


def _build_excel_response(appointments, by_day, date_from, date_to):
    """Genera un archivo .xlsx con dos hojas: Citas y Resumen semanal."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return Response({"detail": "openpyxl no está instalado."}, status=500)

    wb = openpyxl.Workbook()

    # ── Hoja 1: Citas ──
    ws1 = wb.active
    ws1.title = "Citas"

    headers = ["Fecha", "Hora", "Cliente", "Email", "Vehículo", "Servicio", "Estado", "Mecánico", "Trabajo solicitado", "Notas"]
    header_fill = PatternFill(start_color="1D63FF", end_color="1D63FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_labels = {
        "scheduled": "Programada",
        "confirmed": "Confirmada",
        "in_progress": "En progreso",
        "completed": "Completada",
        "cancelled": "Cancelada",
    }

    for row_idx, ap in enumerate(appointments, 2):
        raw_start = ap.get("scheduled_start") or ""
        try:
            from datetime import datetime
            dt = datetime.fromisoformat(raw_start)
            fecha = dt.strftime("%d/%m/%Y")
            hora = dt.strftime("%H:%M")
        except Exception:
            fecha = raw_start
            hora = ""

        vehicle_str = f"{ap.get('vehicle_plate', '')} {ap.get('vehicle_make', '')} {ap.get('vehicle_model', '')}".strip()
        st = ap.get("status") or ""

        ws1.append([
            fecha,
            hora,
            ap.get("customer_name") or "",
            ap.get("customer_email") or "",
            vehicle_str,
            ap.get("service_name") or "",
            status_labels.get(st, st),
            ap.get("mechanic_username") or "",
            ap.get("requested_work") or "",
            ap.get("notes") or "",
        ])

    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Hoja 2: Resumen semanal ──
    ws2 = wb.create_sheet("Resumen semanal")
    ws2.append(["Fecha", "Día", "Cantidad de citas"])
    h2_fill = PatternFill(start_color="0A3EA6", end_color="0A3EA6", fill_type="solid")
    h2_font = Font(bold=True, color="FFFFFF")
    for cell in ws2[1]:
        cell.fill = h2_fill
        cell.font = h2_font

    for row in by_day:
        ws2.append([row["date"], row["day"], row["count"]])

    for col in ws2.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    # Escribir en buffer y devolver
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_citas_{date_from}_al_{date_to}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


class AppointmentCustomerViewSet(viewsets.ModelViewSet):
    serializer_class = AppointmentSerializer
    authentication_classes = [CustomerJWTAuthentication]
    permission_classes = [IsAuthenticatedCustomer]

    def get_queryset(self):
        customer = self.request.user
        return (
            Appointment.objects.select_related("customer", "vehicle", "service", "slot")
            .filter(customer_id=customer.customer_id)
            .order_by("-scheduled_start")
        )

    def create(self, request, *args, **kwargs):
        customer = request.user
        data = request.data or {}

        forbidden = {"scheduled_end", "scheduled_start", "status", "admin_message", "progress_percent", "customer_id"}
        if any(k in data for k in forbidden):
            return Response({"detail": "No permitido."}, status=403)

        vehicle_id = data.get("vehicle_id")
        service_id = data.get("service_id")
        slot_id = data.get("slot_id")
        requested_work = (data.get("requested_work") or "").strip() or None
        notes = (data.get("notes") or "").strip() or None

        if not vehicle_id:
            return Response({"detail": "vehicle_id es requerido."}, status=400)
        if not service_id:
            return Response({"detail": "service_id es requerido."}, status=400)
        if not slot_id:
            return Response({"detail": "slot_id es requerido."}, status=400)

        try:
            vehicle = Vehicle.objects.get(vehicle_id=vehicle_id)
        except Vehicle.DoesNotExist:
            return Response({"detail": "Vehículo no existe."}, status=404)

        if str(vehicle.customer_id) != str(customer.customer_id):
            return Response({"detail": "Ese vehículo no pertenece a tu cuenta."}, status=403)

        try:
            slot = AppointmentSlot.objects.get(slot_id=slot_id, is_active=True)
        except AppointmentSlot.DoesNotExist:
            return Response({"detail": "Horario no disponible."}, status=404)

        used = _count_used_capacity(str(slot.slot_id))
        remaining = max(int(slot.capacity) - used, 0)
        if remaining <= 0:
            return Response({"detail": "No hay cupos disponibles para este horario."}, status=409)

        with connection.cursor() as cursor:
            cursor.execute(
                """
                insert into public.appointments
                  (customer_id, vehicle_id, service_id, slot_id, scheduled_start, scheduled_end,
                   requested_work, status, notes, admin_message, progress_percent, created_at, updated_at)
                values
                  (%s, %s, %s, %s, %s, %s,
                   %s, 'scheduled', %s, null, 0, now(), now())
                returning appointment_id
                """,
                [
                    str(customer.customer_id),
                    str(vehicle.vehicle_id),
                    service_id,
                    str(slot.slot_id),
                    slot.start_at,
                    slot.end_at,
                    requested_work,
                    notes,
                ],
            )
            appointment_id = cursor.fetchone()[0]

        ap = Appointment.objects.select_related("customer", "vehicle", "service", "slot").get(appointment_id=appointment_id)
        return Response(self.get_serializer(ap).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        ap = self.get_object()
        data = request.data or {}

        forbidden = {
            "status",
            "assigned_mechanic_id",
            "created_by",
            "customer_id",
            "scheduled_end",
            "vehicle_id",
            "service_id",
            "slot_id",
            "scheduled_start",
            "admin_message",
            "progress_percent",
        }
        if any(k in data for k in forbidden):
            return Response({"detail": "No permitido."}, status=403)

        sets = []
        params = []

        if "requested_work" in data:
            sets.append("requested_work = %s")
            params.append((data.get("requested_work") or "").strip() or None)

        if "notes" in data:
            sets.append("notes = %s")
            params.append((data.get("notes") or "").strip() or None)

        if not sets:
            ap = Appointment.objects.select_related("customer", "vehicle", "service", "slot").get(appointment_id=ap.appointment_id)
            return Response(self.get_serializer(ap).data, status=200)

        sets.append("updated_at = now()")
        params.append(str(ap.appointment_id))

        with connection.cursor() as cursor:
            cursor.execute(
                f"update public.appointments set {', '.join(sets)} where appointment_id = %s",
                params,
            )

        ap.refresh_from_db()
        ap = Appointment.objects.select_related("customer", "vehicle", "service", "slot").get(appointment_id=ap.appointment_id)
        return Response(self.get_serializer(ap).data, status=200)

    def destroy(self, request, *args, **kwargs):
        ap = self.get_object()

        current_status = (ap.status or "").strip().lower()
        if current_status != "scheduled":
            return Response(
                {"detail": "No podés cancelar una cita que ya fue aceptada o procesada por el taller."},
                status=403,
            )

        with connection.cursor() as cursor:
            cursor.execute(
                """
                update public.appointments
                set status = 'cancelled', updated_at = now()
                where appointment_id = %s and customer_id = %s and status = 'scheduled'
                """,
                [str(ap.appointment_id), str(request.user.customer_id)],
            )

        return Response(status=status.HTTP_204_NO_CONTENT)

    @action(detail=False, methods=["get"], url_path="reminders")
    def reminders(self, request):
        """Devuelve recordatorios de citas próximas (1h y 24h) para el cliente autenticado.
        Excluye citas canceladas. El estado 'visto' se maneja en el frontend (localStorage)."""
        customer = request.user
        now = timezone.now()
        window_end = now + timedelta(hours=25)

        with connection.cursor() as cursor:
            cursor.execute(
                """
                select
                    a.appointment_id,
                    a.scheduled_start,
                    a.status,
                    s.name as service_name,
                    v.plate as vehicle_plate
                from public.appointments a
                left join public.services s on s.service_id = a.service_id
                left join public.vehicles v on v.vehicle_id = a.vehicle_id
                where a.customer_id = %s
                  and a.status != 'cancelled'
                  and a.scheduled_start >= %s
                  and a.scheduled_start <= %s
                order by a.scheduled_start
                """,
                [str(customer.customer_id), now, window_end],
            )
            rows = cursor.fetchall()

        result = []
        for row in rows:
            appt_id, scheduled_start, appt_status, service_name, vehicle_plate = row
            delta = (scheduled_start - now).total_seconds() / 60  # minutos
            if delta <= 60:
                reminder_type = "1h"
            else:
                reminder_type = "24h"
            result.append({
                "appointment_id": str(appt_id),
                "scheduled_start": scheduled_start.isoformat(),
                "status": appt_status,
                "service_name": service_name or "",
                "vehicle_plate": vehicle_plate or "",
                "reminder_type": reminder_type,
            })

        return Response(result, status=200)
