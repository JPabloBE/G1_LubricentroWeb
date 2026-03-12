import calendar
import io
from datetime import date, datetime
from decimal import Decimal

from django.db import connection
from django.http import FileResponse
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.authentication.permissions import IsStaffOrAdmin

from .models import PeriodClosure
from .serializers import PeriodClosureListSerializer, PeriodClosureSerializer


# ── Helpers ────────────────────────────────────────────────────────────────────

def _period_dates(year: int, month: int):
    """Devuelve (period_start, period_end) como objetos date."""
    last_day = calendar.monthrange(year, month)[1]
    return date(year, month, 1), date(year, month, last_day)


def _aggregate_period(period_start: date, period_end: date) -> dict:
    """Agrega datos financieros del periodo desde cash_register (django_app schema)."""
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                COUNT(DISTINCT cs.cash_session_id),
                COUNT(cm.cash_movement_id),
                COALESCE(SUM(CASE WHEN cm.amount > 0 THEN cm.amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN cm.amount < 0 THEN ABS(cm.amount) ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN cm.movement_type = 'sale'       THEN cm.amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN cm.movement_type = 'payment'    THEN cm.amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN cm.movement_type = 'withdrawal' THEN cm.amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN cm.movement_type = 'refund'     THEN cm.amount ELSE 0 END), 0),
                COALESCE(SUM(CASE WHEN cm.movement_type = 'adjustment' THEN cm.amount ELSE 0 END), 0),
                COALESCE(SUM(cc.difference), 0)
            FROM django_app.cash_sessions cs
            LEFT JOIN django_app.cash_movements cm ON cm.cash_session_id = cs.cash_session_id
            LEFT JOIN django_app.cash_closings cc  ON cc.cash_session_id = cs.cash_session_id
            WHERE cs.opened_at::date >= %s AND cs.opened_at::date <= %s
            """,
            [period_start, period_end],
        )
        row = cursor.fetchone()

    total_income = Decimal(str(row[2] or 0))
    total_expenses = Decimal(str(row[3] or 0))
    return {
        "total_sessions": int(row[0] or 0),
        "total_movements": int(row[1] or 0),
        "total_income": total_income,
        "total_expenses": total_expenses,
        "total_net": total_income - total_expenses,
        "sales_total": Decimal(str(row[4] or 0)),
        "payment_total": Decimal(str(row[5] or 0)),
        "withdrawal_total": Decimal(str(row[6] or 0)),
        "refund_total": Decimal(str(row[7] or 0)),
        "adjustment_total": Decimal(str(row[8] or 0)),
        "cash_discrepancies": Decimal(str(row[9] or 0)),
    }


def _open_sessions_in_period(period_start: date, period_end: date) -> list:
    """Devuelve IDs de sesiones aún abiertas en el periodo."""
    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT cash_session_id::text
            FROM django_app.cash_sessions
            WHERE opened_at::date >= %s AND opened_at::date <= %s
              AND status = 'open'
            """,
            [period_start, period_end],
        )
        return [r[0] for r in cursor.fetchall()]


def _insert_audit(closure_id: str, action: str, performed_by: str, notes: str = None):
    with connection.cursor() as cursor:
        cursor.execute(
            """
            INSERT INTO public.period_closure_audit
              (closure_id, action, performed_by, performed_at, notes)
            VALUES (%s, %s, %s, now(), %s)
            """,
            [closure_id, action, performed_by, notes],
        )


# ── PDF Export ─────────────────────────────────────────────────────────────────

def _generate_pdf(closure: PeriodClosure) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=2 * cm, bottomMargin=2 * cm)
    styles = getSampleStyleSheet()
    elements = []

    # Título
    elements.append(Paragraph(f"Cierre Mensual — {closure.folio}", styles["Title"]))
    elements.append(Spacer(1, 0.4 * cm))

    meses_es = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    periodo = f"{meses_es[closure.period_start.month]} {closure.period_start.year}"
    elements.append(Paragraph(f"<b>Periodo:</b> {periodo}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Estado:</b> {closure.status.title()}", styles["Normal"]))
    elements.append(Paragraph(
        f"<b>Cerrado el:</b> {closure.closed_at.strftime('%d/%m/%Y %H:%M') if closure.closed_at else '—'}",
        styles["Normal"],
    ))
    elements.append(Spacer(1, 0.6 * cm))

    # Resumen financiero
    elements.append(Paragraph("<b>Resumen Financiero</b>", styles["Heading2"]))
    summary_data = [
        ["Concepto", "Monto"],
        ["Ingresos totales", f"${closure.total_income:,.2f}"],
        ["Egresos totales", f"${closure.total_expenses:,.2f}"],
        ["Utilidad neta", f"${closure.total_net:,.2f}"],
        ["Sesiones de caja", str(closure.total_sessions)],
        ["Total movimientos", str(closure.total_movements)],
        ["Diferencias de caja", f"${closure.cash_discrepancies:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[10 * cm, 6 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#212529")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        ("FONTNAME", (0, -3), (-1, -3), "Helvetica-Bold"),  # Utilidad neta en negrita
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.6 * cm))

    # Desglose por tipo
    elements.append(Paragraph("<b>Desglose por Tipo de Movimiento</b>", styles["Heading2"]))
    type_labels = {
        "sales_total": "Ventas",
        "payment_total": "Pagos",
        "withdrawal_total": "Retiros",
        "refund_total": "Devoluciones",
        "adjustment_total": "Ajustes",
    }
    breakdown_data = [["Tipo", "Monto"]]
    for field, label in type_labels.items():
        val = getattr(closure, field, Decimal("0"))
        breakdown_data.append([label, f"${val:,.2f}"])

    breakdown_table = Table(breakdown_data, colWidths=[10 * cm, 6 * cm])
    breakdown_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#495057")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9fa")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dee2e6")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(breakdown_table)

    if closure.notes:
        elements.append(Spacer(1, 0.6 * cm))
        elements.append(Paragraph("<b>Notas</b>", styles["Heading2"]))
        elements.append(Paragraph(closure.notes, styles["Normal"]))

    doc.build(elements)
    return buffer.getvalue()


# ── Excel Export ───────────────────────────────────────────────────────────────

def _generate_excel(closure: PeriodClosure) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()

    # Hoja 1: Resumen
    ws = wb.active
    ws.title = "Resumen"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="212529")
    meses_es = ["", "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    periodo = f"{meses_es[closure.period_start.month]} {closure.period_start.year}"

    ws.append([f"Cierre Mensual — {closure.folio}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Periodo: {periodo}"])
    ws.append([f"Estado: {closure.status.title()}"])
    ws.append([])

    ws.append(["Concepto", "Monto"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill

    rows = [
        ("Ingresos totales", float(closure.total_income)),
        ("Egresos totales", float(closure.total_expenses)),
        ("Utilidad neta", float(closure.total_net)),
        ("Sesiones de caja", closure.total_sessions),
        ("Total movimientos", closure.total_movements),
        ("Diferencias de caja", float(closure.cash_discrepancies)),
        (None, None),
        ("— Desglose por tipo —", None),
        ("Ventas", float(closure.sales_total)),
        ("Pagos", float(closure.payment_total)),
        ("Retiros", float(closure.withdrawal_total)),
        ("Devoluciones", float(closure.refund_total)),
        ("Ajustes", float(closure.adjustment_total)),
    ]
    for row in rows:
        ws.append(row)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 18

    # Hoja 2: Movimientos del periodo
    ws2 = wb.create_sheet("Movimientos")
    ws2.append(["Sesión", "Tipo", "Descripción", "Monto", "OT vinculada", "Fecha"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill

    with connection.cursor() as cursor:
        cursor.execute(
            """
            SELECT
                cm.cash_session_id::text,
                cm.movement_type,
                cm.description,
                cm.amount,
                cm.work_order_id::text,
                cm.created_at
            FROM django_app.cash_movements cm
            JOIN django_app.cash_sessions cs ON cs.cash_session_id = cm.cash_session_id
            WHERE cs.opened_at::date >= %s AND cs.opened_at::date <= %s
            ORDER BY cm.created_at ASC
            """,
            [closure.period_start, closure.period_end],
        )
        for r in cursor.fetchall():
            session_short = str(r[0])[:8] if r[0] else "—"
            created_at = r[5].strftime("%d/%m/%Y %H:%M") if r[5] else "—"
            ws2.append([session_short, r[1], r[2] or "", float(r[3] or 0), r[4] or "—", created_at])

    for col in ["A", "B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── ViewSet ────────────────────────────────────────────────────────────────────

class PeriodClosureViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]
    http_method_names = ["get", "post", "head", "options"]

    def get_serializer_class(self):
        if self.action == "list":
            return PeriodClosureListSerializer
        return PeriodClosureSerializer

    def get_queryset(self):
        return PeriodClosure.objects.prefetch_related("audit_entries").all()

    # ── POST /api/period-closures/preview/ ──────────────────────────────────
    @action(detail=False, methods=["post"], url_path="preview")
    def preview(self, request):
        year = request.data.get("year")
        month = request.data.get("month")
        if not year or not month:
            return Response({"detail": "year y month son requeridos."}, status=400)
        try:
            year, month = int(year), int(month)
            if not (1 <= month <= 12):
                raise ValueError
        except (ValueError, TypeError):
            return Response({"detail": "year/month inválidos."}, status=400)

        period_start, period_end = _period_dates(year, month)
        folio = f"CM-{year}-{month:02d}"

        # Verificar cierre existente
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT closure_id FROM public.period_closures WHERE folio = %s LIMIT 1",
                [folio],
            )
            existing = cursor.fetchone()

        if existing:
            return Response({"detail": f"Ya existe un cierre para {folio}.", "folio": folio}, status=409)

        # Sesiones abiertas en el periodo
        open_sessions = _open_sessions_in_period(period_start, period_end)
        aggregated = _aggregate_period(period_start, period_end)

        warnings = []
        if open_sessions:
            warnings.append(f"Hay {len(open_sessions)} sesión(es) de caja aún abiertas en este periodo.")
        if aggregated["cash_discrepancies"] != 0:
            warnings.append(f"Diferencias de caja detectadas: ${aggregated['cash_discrepancies']:,.2f}")

        return Response({
            "folio": folio,
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "aggregated": {k: str(v) for k, v in aggregated.items()},
            "open_sessions": open_sessions,
            "warnings": warnings,
            "can_close": len(open_sessions) == 0,
        })

    # ── POST /api/period-closures/ ──────────────────────────────────────────
    def create(self, request, *args, **kwargs):
        year = request.data.get("year")
        month = request.data.get("month")
        notes = (request.data.get("notes") or "").strip() or None
        force = request.data.get("force", False)

        if not year or not month:
            return Response({"detail": "year y month son requeridos."}, status=400)
        try:
            year, month = int(year), int(month)
            if not (1 <= month <= 12):
                raise ValueError
        except (ValueError, TypeError):
            return Response({"detail": "year/month inválidos."}, status=400)

        period_start, period_end = _period_dates(year, month)
        folio = f"CM-{year}-{month:02d}"

        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT closure_id FROM public.period_closures WHERE folio = %s LIMIT 1", [folio]
            )
            if cursor.fetchone():
                return Response({"detail": f"Ya existe un cierre para {folio}."}, status=409)

        open_sessions = _open_sessions_in_period(period_start, period_end)
        if open_sessions and not force:
            return Response({
                "detail": "Hay sesiones de caja abiertas. Use force=true para cerrar de igual manera.",
                "open_sessions": open_sessions,
            }, status=409)

        agg = _aggregate_period(period_start, period_end)
        user_id = str(request.user.id)

        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO public.period_closures
                  (closure_type, period_start, period_end, folio, status,
                   total_income, total_expenses, total_net,
                   total_sessions, total_movements, cash_discrepancies,
                   sales_total, payment_total, withdrawal_total, refund_total, adjustment_total,
                   notes, closed_by, closed_at, created_at, updated_at)
                VALUES
                  ('monthly', %s, %s, %s, 'closed',
                   %s, %s, %s,
                   %s, %s, %s,
                   %s, %s, %s, %s, %s,
                   %s, %s, now(), now(), now())
                RETURNING closure_id
                """,
                [
                    period_start, period_end, folio,
                    str(agg["total_income"]), str(agg["total_expenses"]), str(agg["total_net"]),
                    agg["total_sessions"], agg["total_movements"], str(agg["cash_discrepancies"]),
                    str(agg["sales_total"]), str(agg["payment_total"]),
                    str(agg["withdrawal_total"]), str(agg["refund_total"]),
                    str(agg["adjustment_total"]),
                    notes, user_id,
                ],
            )
            closure_id = str(cursor.fetchone()[0])

        _insert_audit(closure_id, "created", user_id, f"Cierre mensual {folio} creado.")

        closure = PeriodClosure.objects.prefetch_related("audit_entries").get(closure_id=closure_id)
        return Response(PeriodClosureSerializer(closure).data, status=status.HTTP_201_CREATED)

    # ── POST /api/period-closures/{id}/reopen/ ──────────────────────────────
    @action(detail=True, methods=["post"], url_path="reopen")
    def reopen(self, request, pk=None):
        if request.user.user_type != "admin":
            return Response({"detail": "Solo administradores pueden reabrir cierres."}, status=403)

        closure = self.get_object()
        reopen_reason = (request.data.get("reopen_reason") or "").strip()
        if not reopen_reason:
            return Response({"detail": "reopen_reason es requerido."}, status=400)

        user_id = str(request.user.id)
        with connection.cursor() as cursor:
            cursor.execute(
                """
                UPDATE public.period_closures
                SET status = 'reopened',
                    reopened_by = %s,
                    reopened_at = now(),
                    reopen_reason = %s,
                    updated_at = now()
                WHERE closure_id = %s
                """,
                [user_id, reopen_reason, str(closure.closure_id)],
            )

        _insert_audit(str(closure.closure_id), "reopened", user_id, reopen_reason)

        closure = PeriodClosure.objects.prefetch_related("audit_entries").get(closure_id=closure.closure_id)
        return Response(PeriodClosureSerializer(closure).data)

    # ── GET /api/period-closures/{id}/export/pdf/ ───────────────────────────
    @action(detail=True, methods=["get"], url_path="export/pdf")
    def export_pdf(self, request, pk=None):
        closure = self.get_object()
        pdf_bytes = _generate_pdf(closure)
        _insert_audit(str(closure.closure_id), "exported_pdf", str(request.user.id))

        response = FileResponse(
            io.BytesIO(pdf_bytes),
            content_type="application/pdf",
            as_attachment=True,
            filename=f"{closure.folio}.pdf",
        )
        return response

    # ── GET /api/period-closures/{id}/export/excel/ ─────────────────────────
    @action(detail=True, methods=["get"], url_path="export/excel")
    def export_excel(self, request, pk=None):
        closure = self.get_object()
        excel_bytes = _generate_excel(closure)
        _insert_audit(str(closure.closure_id), "exported_excel", str(request.user.id))

        response = FileResponse(
            io.BytesIO(excel_bytes),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            filename=f"{closure.folio}.xlsx",
        )
        return response
