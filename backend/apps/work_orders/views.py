import io
from decimal import Decimal, InvalidOperation

from django.db import connection, transaction
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from apps.authentication.permissions import IsStaffOrAdmin
from apps.catalog.stock import log_stock_movement
from apps.customers.auth import CustomerJWTAuthentication
from apps.customers.permissions import IsAuthenticatedCustomer

from .models import WorkOrder, WorkOrderProduct, WorkOrderService
from .serializers import (
    WorkOrderCustomerSerializer,
    WorkOrderProductSerializer,
    WorkOrderSerializer,
    WorkOrderServiceSerializer,
)

PRODUCT_STOCK_COLUMN = "stock_qty"


def _to_decimal(value, field_name: str) -> Decimal:
    try:
        d = Decimal(str(value))
    except (InvalidOperation, TypeError):
        raise ValueError(f"{field_name} inválido.")
    if d <= 0:
        raise ValueError(f"{field_name} debe ser > 0.")
    return d


def _ensure_work_order_not_cancelled(work_order_id: str) -> None:
    with connection.cursor() as cursor:
        cursor.execute("select status from public.work_orders where work_order_id = %s", [work_order_id])
        row = cursor.fetchone()
        if not row:
            raise ValueError("Work order no existe.")
        if (row[0] or "").strip().lower() == "cancelled":
            raise ValueError("No se puede modificar una work order cancelada.")


def _lock_product_and_get_stock(product_id: str) -> Decimal:
    with connection.cursor() as cursor:
        cursor.execute(
            f"select {PRODUCT_STOCK_COLUMN} from public.products where product_id = %s for update",
            [product_id],
        )
        row = cursor.fetchone()
        if not row:
            raise ValueError("Producto no existe.")
        return Decimal(str(row[0] or 0))


def _update_product_stock(product_id: str, new_stock: Decimal) -> None:
    with connection.cursor() as cursor:
        cursor.execute(
            f"""
            update public.products
            set {PRODUCT_STOCK_COLUMN} = %s,
                updated_at = now()
            where product_id = %s
            """,
            [str(new_stock), product_id],
        )


class OpenAppointmentsAdminViewSet(viewsets.ViewSet):
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]

    def list(self, request):
        statuses = request.query_params.get("statuses")
        if statuses:
            allowed = [s.strip() for s in statuses.split(",") if s.strip()]
        else:
            allowed = ["pending", "scheduled", "accepted", "in_progress"]

        with connection.cursor() as cursor:
            cursor.execute(
                """
                select
                  a.appointment_id,
                  a.status,
                  a.scheduled_start,
                  a.requested_work,
                  c.customer_id,
                  c.full_name,
                  c.email,
                  v.vehicle_id,
                  v.plate,
                  v.make,
                  v.model,
                  v.year
                from public.appointments a
                join public.customers c on c.customer_id = a.customer_id
                join public.vehicles v on v.vehicle_id = a.vehicle_id
                where a.status = any(%s)
                order by a.scheduled_start asc
                limit 200
                """,
                [allowed],
            )
            rows = cursor.fetchall()

        data = []
        for r in rows:
            data.append(
                {
                    "appointment_id": r[0],
                    "status": r[1],
                    "scheduled_start": r[2],
                    "requested_work": r[3],
                    "customer": {"customer_id": r[4], "full_name": r[5], "email": r[6]},
                    "vehicle": {"vehicle_id": r[7], "plate": r[8], "make": r[9], "model": r[10], "year": r[11]},
                }
            )

        return Response(data, status=200)


class WorkOrderAdminViewSet(viewsets.ModelViewSet):
    serializer_class = WorkOrderSerializer
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]

    def get_queryset(self):
        qs = WorkOrder.objects.select_related("customer", "vehicle", "appointment").all()

        appointment_id = self.request.query_params.get("appointment_id")
        if appointment_id:
            qs = qs.filter(appointment_id=appointment_id)

        customer_id = self.request.query_params.get("customer_id")
        if customer_id:
            qs = qs.filter(customer_id=customer_id)

        status_q = self.request.query_params.get("status")
        if status_q:
            qs = qs.filter(status__iexact=status_q.strip())

        statuses_q = self.request.query_params.get("statuses")
        if statuses_q:
            allowed = [s.strip().lower() for s in statuses_q.split(",") if s.strip()]
            if allowed:
                qs = qs.filter(status__in=allowed)

        date_from = self.request.query_params.get("date_from")
        if date_from:
            qs = qs.filter(opened_at__date__gte=date_from)

        date_to = self.request.query_params.get("date_to")
        if date_to:
            qs = qs.filter(opened_at__date__lte=date_to)

        mechanic_id = self.request.query_params.get("mechanic_id")
        if mechanic_id:
            qs = qs.filter(assigned_mechanic_id=mechanic_id)

        return qs.order_by("-opened_at")

    @action(detail=False, methods=["get"], url_path="report")
    def report(self, request):
        """Reporte de órdenes de trabajo con filtros por fecha, estado y mecánico.
        Soporta export=excel para descargar .xlsx (openpyxl)."""
        params = request.query_params
        today = timezone.now().date()
        month_start = today.replace(day=1)

        date_from = (params.get("date_from") or str(month_start)).strip()
        date_to = (params.get("date_to") or str(today)).strip()
        status_filter = (params.get("status") or "").strip()
        mechanic_id_filter = (params.get("mechanic_id") or "").strip()
        export_format = (params.get("export") or "json").strip().lower()

        filters = ["wo.opened_at::date >= %s", "wo.opened_at::date <= %s"]
        values = [date_from, date_to]

        if status_filter:
            filters.append("wo.status = %s")
            values.append(status_filter)

        if mechanic_id_filter:
            filters.append("wo.assigned_mechanic_id::text = %s")
            values.append(mechanic_id_filter)

        where_clause = " and ".join(filters)

        with connection.cursor() as cursor:
            cursor.execute(
                f"""
                select
                    wo.work_order_id,
                    wo.status,
                    wo.authorization_status,
                    wo.estimated_total,
                    wo.opened_at,
                    wo.closed_at,
                    wo.customer_symptoms,
                    wo.diagnosis,
                    wo.notes,
                    c.full_name                         as customer_name,
                    c.email                             as customer_email,
                    v.plate                             as vehicle_plate,
                    v.make                              as vehicle_make,
                    v.model                             as vehicle_model,
                    v.year                              as vehicle_year,
                    u.first_name || ' ' || u.last_name  as mechanic_name,
                    u.username                          as mechanic_username
                from public.work_orders wo
                left join public.customers      c on c.customer_id = wo.customer_id
                left join public.vehicles       v on v.vehicle_id  = wo.vehicle_id
                left join django_app.auth_users u on u.id = wo.assigned_mechanic_id
                where {where_clause}
                order by wo.opened_at desc
                """,
                values,
            )
            rows = cursor.fetchall()
            cols = [d[0] for d in cursor.description]

        work_orders = [dict(zip(cols, row)) for row in rows]

        # Serializar campos no-JSON (datetimes, Decimals, UUIDs)
        for wo in work_orders:
            for k, v in wo.items():
                if hasattr(v, "isoformat"):
                    wo[k] = v.isoformat()
                elif v is None:
                    wo[k] = None
                elif not isinstance(v, (str, int, float, bool)):
                    wo[k] = str(v)

        # Resumen por estado
        status_keys = ["open", "in_progress", "ready", "closed", "cancelled"]
        by_status = {k: 0 for k in status_keys}
        for wo in work_orders:
            st = wo.get("status") or "unknown"
            if st in by_status:
                by_status[st] += 1

        total_estimated = sum(
            Decimal(str(wo["estimated_total"]))
            for wo in work_orders
            if wo.get("estimated_total") is not None
        )

        summary = {
            "total": len(work_orders),
            "by_status": by_status,
            "total_estimated": str(total_estimated),
        }

        if export_format == "excel":
            return _build_wo_excel_response(work_orders, summary, date_from, date_to)

        return Response({"work_orders": work_orders, "summary": summary}, status=200)

    @action(detail=False, methods=["post"], url_path="create-from-appointment")
    def create_from_appointment(self, request):
        appointment_id = (request.data or {}).get("appointment_id")
        if not appointment_id:
            return Response({"detail": "appointment_id es requerido."}, status=400)

        with connection.cursor() as cursor:
            cursor.execute(
                "select customer_id, vehicle_id from public.appointments where appointment_id = %s",
                [str(appointment_id)],
            )
            row = cursor.fetchone()
            if not row:
                return Response({"detail": "Cita no existe."}, status=404)

            customer_id, vehicle_id = row[0], row[1]

            cursor.execute(
                "select work_order_id from public.work_orders where appointment_id = %s",
                [str(appointment_id)],
            )
            exists = cursor.fetchone()
            if exists:
                wo_id = exists[0]
                wo = WorkOrder.objects.select_related("customer", "vehicle", "appointment").get(work_order_id=wo_id)
                return Response(self.get_serializer(wo).data, status=200)

            cursor.execute(
                """
                insert into public.work_orders
                  (appointment_id, customer_id, vehicle_id, status, authorization_status, opened_at, created_at, updated_at)
                values
                  (%s, %s, %s, 'open', 'pending', now(), now(), now())
                returning work_order_id
                """,
                [str(appointment_id), str(customer_id), str(vehicle_id)],
            )
            wo_id = cursor.fetchone()[0]

        wo = WorkOrder.objects.select_related("customer", "vehicle", "appointment").get(work_order_id=wo_id)
        return Response(self.get_serializer(wo).data, status=201)

    def create(self, request, *args, **kwargs):
        data = request.data or {}
        customer_id = data.get("customer_id")
        vehicle_id = data.get("vehicle_id")
        if not customer_id:
            return Response({"detail": "customer_id es requerido."}, status=400)
        if not vehicle_id:
            return Response({"detail": "vehicle_id es requerido."}, status=400)

        appointment_id = data.get("appointment_id") or None
        status_v = (data.get("status") or "open").strip()
        auth_status = (data.get("authorization_status") or "pending").strip()
        customer_symptoms = (data.get("customer_symptoms") or "").strip() or None
        notes = (data.get("notes") or "").strip() or None

        with connection.cursor() as cursor:
            cursor.execute(
                """
                insert into public.work_orders
                  (appointment_id, customer_id, vehicle_id, status, customer_symptoms,
                   authorization_status, opened_at, notes, created_at, updated_at)
                values
                  (%s, %s, %s, %s, %s,
                   %s, now(), %s, now(), now())
                returning work_order_id
                """,
                [appointment_id, customer_id, vehicle_id, status_v, customer_symptoms, auth_status, notes],
            )
            work_order_id = cursor.fetchone()[0]

        wo = WorkOrder.objects.select_related("customer", "vehicle", "appointment").get(work_order_id=work_order_id)
        return Response(self.get_serializer(wo).data, status=status.HTTP_201_CREATED)

    def partial_update(self, request, *args, **kwargs):
        wo = self.get_object()
        data = request.data or {}

        allowed = {
            "status",
            "customer_symptoms",
            "diagnosis",
            "estimated_total",
            "authorization_status",
            "authorized_at",
            "authorized_by",
            "assigned_mechanic_id",
            "closed_at",
            "notes",
        }
        for k in data.keys():
            if k not in allowed:
                return Response({"detail": f"No permitido: {k}"}, status=403)

        sets = []
        params = []

        if "status" in data:
            sets.append("status = %s")
            params.append((data.get("status") or "").strip())

        if "customer_symptoms" in data:
            sets.append("customer_symptoms = %s")
            params.append((data.get("customer_symptoms") or "").strip() or None)

        if "diagnosis" in data:
            sets.append("diagnosis = %s")
            params.append((data.get("diagnosis") or "").strip() or None)

        if "estimated_total" in data:
            sets.append("estimated_total = %s")
            params.append(data.get("estimated_total"))

        if "authorization_status" in data:
            auth_val = (data.get("authorization_status") or "").strip()
            if auth_val not in ("pending", "approved", "rejected"):
                return Response(
                    {"detail": "authorization_status inválido. Valores permitidos: pending, approved, rejected."},
                    status=400,
                )
            sets.append("authorization_status = %s")
            params.append(auth_val)

        if "authorized_at" in data:
            sets.append("authorized_at = %s")
            params.append(data.get("authorized_at") or None)

        if "authorized_by" in data:
            sets.append("authorized_by = %s")
            params.append((data.get("authorized_by") or "").strip() or None)

        if "assigned_mechanic_id" in data:
            sets.append("assigned_mechanic_id = %s")
            params.append(data.get("assigned_mechanic_id") or None)

        if "closed_at" in data:
            sets.append("closed_at = %s")
            params.append(data.get("closed_at") or None)

        if "notes" in data:
            sets.append("notes = %s")
            params.append((data.get("notes") or "").strip() or None)

        if not sets:
            wo.refresh_from_db()
            return Response(self.get_serializer(wo).data, status=200)

        sets.append("updated_at = now()")
        params.append(str(wo.work_order_id))

        with connection.cursor() as cursor:
            cursor.execute(
                f"update public.work_orders set {', '.join(sets)} where work_order_id = %s",
                params,
            )

        wo.refresh_from_db()
        wo = WorkOrder.objects.select_related("customer", "vehicle", "appointment").get(work_order_id=wo.work_order_id)
        return Response(self.get_serializer(wo).data, status=200)

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        wo = self.get_object()
        wo_id = str(wo.work_order_id)

        with connection.cursor() as cursor:
            cursor.execute(
                "select work_order_id from public.work_orders where work_order_id = %s for update",
                [wo_id],
            )
            locked = cursor.fetchone()
            if not locked:
                return Response({"detail": "Work order no existe."}, status=404)

            cursor.execute(
                "select product_id, qty from public.work_order_products where work_order_id = %s",
                [wo_id],
            )
            prod_lines = cursor.fetchall() or []

            for product_id, qty in prod_lines:
                if not product_id:
                    continue

                cursor.execute(
                    f"select {PRODUCT_STOCK_COLUMN} from public.products where product_id = %s for update",
                    [str(product_id)],
                )
                pr = cursor.fetchone()
                if pr is None:
                    continue

                current_stock = Decimal(str(pr[0] or 0))
                qty_dec = Decimal(str(qty))
                new_stock = current_stock + qty_dec
                cursor.execute(
                    f"""
                    update public.products
                    set {PRODUCT_STOCK_COLUMN} = %s,
                        updated_at = now()
                    where product_id = %s
                    """,
                    [str(new_stock), str(product_id)],
                )
                log_stock_movement(
                    product_id=str(product_id),
                    qty_before=current_stock,
                    qty_change=qty_dec,
                    qty_after=new_stock,
                    movement_type="work_order_refund",
                    reason=f"Cancelación OT {wo_id}",
                )

            cursor.execute("delete from public.work_order_products where work_order_id = %s", [wo_id])
            cursor.execute("delete from public.work_order_services where work_order_id = %s", [wo_id])
            cursor.execute("delete from public.work_orders where work_order_id = %s", [wo_id])

        return Response(status=status.HTTP_204_NO_CONTENT)


class WorkOrderCustomerViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = WorkOrderCustomerSerializer
    authentication_classes = [CustomerJWTAuthentication]
    permission_classes = [IsAuthenticatedCustomer]

    def get_queryset(self):
        customer = self.request.user
        return (
            WorkOrder.objects.select_related("vehicle", "appointment")
            .prefetch_related("service_lines__service", "product_lines__product")
            .filter(customer_id=customer.customer_id)
            .order_by("-created_at")
        )


class WorkOrderServiceAdminViewSet(viewsets.ModelViewSet):
    serializer_class = WorkOrderServiceSerializer
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]

    def get_queryset(self):
        qs = WorkOrderService.objects.select_related("work_order", "service").all()
        work_order_id = self.request.query_params.get("work_order_id")
        if work_order_id:
            qs = qs.filter(work_order_id=work_order_id)
        return qs

    def create(self, request, *args, **kwargs):
        data = request.data or {}
        work_order_id = data.get("work_order_id")
        if not work_order_id:
            return Response({"detail": "work_order_id es requerido."}, status=400)

        try:
            _ensure_work_order_not_cancelled(str(work_order_id))
            qty = _to_decimal(data.get("qty", "1"), "qty")
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        unit_price_raw = data.get("unit_price", "0")
        try:
            unit_price = Decimal(str(unit_price_raw))
        except Exception:
            return Response({"detail": "unit_price inválido."}, status=400)
        if unit_price < 0:
            return Response({"detail": "unit_price no puede ser negativo."}, status=400)

        service_id = data.get("service_id") or None
        desc = (data.get("description") or "").strip() or None
        mechanic_id = data.get("mechanic_id") or None
        status_v = (data.get("status") or "pending").strip()
        started_at = data.get("started_at") or None
        completed_at = data.get("completed_at") or None

        with connection.cursor() as cursor:
            cursor.execute(
                """
                insert into public.work_order_services
                  (work_order_id, service_id, description, qty, unit_price, mechanic_id, status,
                   started_at, completed_at, created_at, updated_at)
                values
                  (%s, %s, %s, %s, %s, %s, %s,
                   %s, %s, now(), now())
                returning work_order_service_id
                """,
                [work_order_id, service_id, desc, str(qty), str(unit_price), mechanic_id, status_v, started_at, completed_at],
            )
            line_id = cursor.fetchone()[0]

        line = WorkOrderService.objects.select_related("work_order", "service").get(work_order_service_id=line_id)
        return Response(self.get_serializer(line).data, status=201)


class WorkOrderProductAdminViewSet(viewsets.ModelViewSet):
    serializer_class = WorkOrderProductSerializer
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]

    def get_queryset(self):
        qs = WorkOrderProduct.objects.select_related("work_order", "product").all()
        work_order_id = self.request.query_params.get("work_order_id")
        if work_order_id:
            qs = qs.filter(work_order_id=work_order_id)
        return qs

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        data = request.data or {}
        work_order_id = data.get("work_order_id")
        product_id = data.get("product_id")

        if not work_order_id:
            return Response({"detail": "work_order_id es requerido."}, status=400)
        if not product_id:
            return Response({"detail": "product_id es requerido."}, status=400)

        try:
            _ensure_work_order_not_cancelled(str(work_order_id))
            qty = _to_decimal(data.get("qty", "1"), "qty")
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        unit_price_raw = data.get("unit_price", "0")
        try:
            unit_price = Decimal(str(unit_price_raw))
        except Exception:
            return Response({"detail": "unit_price inválido."}, status=400)
        if unit_price < 0:
            return Response({"detail": "unit_price no puede ser negativo."}, status=400)

        desc = (data.get("description") or "").strip() or None

        try:
            current_stock = _lock_product_and_get_stock(str(product_id))
        except ValueError as e:
            return Response({"detail": str(e)}, status=404)

        if current_stock < qty:
            return Response({"detail": "Stock insuficiente para este producto."}, status=409)

        new_stock = current_stock - qty
        _update_product_stock(str(product_id), new_stock)

        with connection.cursor() as cursor:
            cursor.execute(
                """
                insert into public.work_order_products
                  (work_order_id, product_id, description, qty, unit_price, created_at, updated_at)
                values
                  (%s, %s, %s, %s, %s, now(), now())
                returning work_order_product_id
                """,
                [work_order_id, product_id, desc, str(qty), str(unit_price)],
            )
            line_id = cursor.fetchone()[0]

        log_stock_movement(
            product_id=str(product_id),
            qty_before=current_stock,
            qty_change=-qty,
            qty_after=new_stock,
            movement_type="work_order",
            performed_by=request.user.id,
            reason=f"OT {work_order_id}",
            reference_id=line_id,
            reference_type="work_order_product",
        )

        line = WorkOrderProduct.objects.select_related("work_order", "product").get(work_order_product_id=line_id)
        return Response(self.get_serializer(line).data, status=201)

    @transaction.atomic
    def partial_update(self, request, *args, **kwargs):
        line = self.get_object()
        data = request.data or {}

        try:
            _ensure_work_order_not_cancelled(str(line.work_order_id))
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        sets = []
        params = []

        if "description" in data:
            sets.append("description = %s")
            params.append((data.get("description") or "").strip() or None)

        if "unit_price" in data:
            unit_price_raw = data.get("unit_price", "0")
            try:
                unit_price = Decimal(str(unit_price_raw))
            except Exception:
                return Response({"detail": "unit_price inválido."}, status=400)
            if unit_price < 0:
                return Response({"detail": "unit_price no puede ser negativo."}, status=400)
            sets.append("unit_price = %s")
            params.append(str(unit_price))

        if "qty" in data:
            try:
                new_qty = _to_decimal(data.get("qty"), "qty")
            except ValueError as e:
                return Response({"detail": str(e)}, status=400)

            old_qty = Decimal(str(line.qty))
            delta = new_qty - old_qty

            if line.product_id:
                try:
                    current_stock = _lock_product_and_get_stock(str(line.product_id))
                except ValueError as e:
                    return Response({"detail": str(e)}, status=404)

                if delta > 0:
                    if current_stock < delta:
                        return Response({"detail": "Stock insuficiente para aumentar la cantidad."}, status=409)
                    new_stock = current_stock - delta
                    _update_product_stock(str(line.product_id), new_stock)
                    log_stock_movement(
                        product_id=str(line.product_id),
                        qty_before=current_stock,
                        qty_change=-delta,
                        qty_after=new_stock,
                        movement_type="work_order",
                        performed_by=request.user.id,
                        reason=f"Ajuste qty OT {line.work_order_id}",
                        reference_id=line.work_order_product_id,
                        reference_type="work_order_product",
                    )
                elif delta < 0:
                    new_stock = current_stock + (-delta)
                    _update_product_stock(str(line.product_id), new_stock)
                    log_stock_movement(
                        product_id=str(line.product_id),
                        qty_before=current_stock,
                        qty_change=-delta,
                        qty_after=new_stock,
                        movement_type="work_order_refund",
                        performed_by=request.user.id,
                        reason=f"Ajuste qty OT {line.work_order_id}",
                        reference_id=line.work_order_product_id,
                        reference_type="work_order_product",
                    )

            sets.append("qty = %s")
            params.append(str(new_qty))

        if not sets:
            line.refresh_from_db()
            return Response(self.get_serializer(line).data, status=200)

        sets.append("updated_at = now()")
        params.append(str(line.work_order_product_id))

        with connection.cursor() as cursor:
            cursor.execute(
                f"update public.work_order_products set {', '.join(sets)} where work_order_product_id = %s",
                params,
            )

        line.refresh_from_db()
        line = WorkOrderProduct.objects.select_related("work_order", "product").get(work_order_product_id=line.work_order_product_id)
        return Response(self.get_serializer(line).data, status=200)

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        line = self.get_object()

        try:
            _ensure_work_order_not_cancelled(str(line.work_order_id))
        except ValueError as e:
            return Response({"detail": str(e)}, status=400)

        if line.product_id:
            try:
                current_stock = _lock_product_and_get_stock(str(line.product_id))
            except ValueError as e:
                return Response({"detail": str(e)}, status=404)
            qty_dec = Decimal(str(line.qty))
            new_stock = current_stock + qty_dec
            _update_product_stock(str(line.product_id), new_stock)
            log_stock_movement(
                product_id=str(line.product_id),
                qty_before=current_stock,
                qty_change=qty_dec,
                qty_after=new_stock,
                movement_type="work_order_refund",
                performed_by=request.user.id,
                reason=f"Eliminación de línea OT {line.work_order_id}",
                reference_id=line.work_order_product_id,
                reference_type="work_order_product",
            )

        with connection.cursor() as cursor:
            cursor.execute(
                "delete from public.work_order_products where work_order_product_id = %s",
                [str(line.work_order_product_id)],
            )

        return Response(status=status.HTTP_204_NO_CONTENT)


def _build_wo_excel_response(work_orders, summary, date_from, date_to):
    """Genera un archivo .xlsx con detalle y resumen de órdenes de trabajo."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        return HttpResponse("openpyxl no está instalado.", status=500)

    STATUS_LABELS = {
        "open": "Abierta",
        "in_progress": "En proceso",
        "ready": "Lista",
        "closed": "Cerrada",
        "cancelled": "Cancelada",
    }
    AUTH_LABELS = {
        "pending": "Pendiente",
        "approved": "Aprobada",
        "rejected": "Rechazada",
    }

    wb = openpyxl.Workbook()

    # ── Hoja 1: Órdenes de Trabajo ──
    ws1 = wb.active
    ws1.title = "Órdenes de Trabajo"

    headers = [
        "Fecha apertura", "Hora apertura", "Fecha cierre",
        "Cliente", "Email",
        "Vehículo", "Año",
        "Estado", "Mecánico",
        "Total estimado", "Autorización",
        "Síntomas", "Diagnóstico", "Notas",
    ]
    header_fill = PatternFill(start_color="1D63FF", end_color="1D63FF", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for wo in work_orders:
        from datetime import datetime

        raw_opened = wo.get("opened_at") or ""
        try:
            dt_open = datetime.fromisoformat(raw_opened)
            fecha_open = dt_open.strftime("%d/%m/%Y")
            hora_open = dt_open.strftime("%H:%M")
        except Exception:
            fecha_open = raw_opened
            hora_open = ""

        raw_closed = wo.get("closed_at") or ""
        try:
            fecha_close = datetime.fromisoformat(raw_closed).strftime("%d/%m/%Y") if raw_closed else ""
        except Exception:
            fecha_close = raw_closed

        vehicle_str = f"{wo.get('vehicle_plate', '')} {wo.get('vehicle_make', '')} {wo.get('vehicle_model', '')}".strip()
        st = wo.get("status") or ""
        auth_st = wo.get("authorization_status") or ""

        estimated = wo.get("estimated_total")
        try:
            estimated_val = float(estimated) if estimated is not None else ""
        except Exception:
            estimated_val = estimated or ""

        ws1.append([
            fecha_open,
            hora_open,
            fecha_close,
            wo.get("customer_name") or "",
            wo.get("customer_email") or "",
            vehicle_str,
            wo.get("vehicle_year") or "",
            STATUS_LABELS.get(st, st),
            wo.get("mechanic_name") or wo.get("mechanic_username") or "",
            estimated_val,
            AUTH_LABELS.get(auth_st, auth_st),
            wo.get("customer_symptoms") or "",
            wo.get("diagnosis") or "",
            wo.get("notes") or "",
        ])

    for col in ws1.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    # ── Hoja 2: Resumen ──
    ws2 = wb.create_sheet("Resumen")
    ws2.append(["Estado", "Cantidad"])
    h2_fill = PatternFill(start_color="0A3EA6", end_color="0A3EA6", fill_type="solid")
    h2_font = Font(bold=True, color="FFFFFF")
    for cell in ws2[1]:
        cell.fill = h2_fill
        cell.font = h2_font

    by_status = summary.get("by_status", {})
    for st_key, label in [
        ("open", "Abierta"),
        ("in_progress", "En proceso"),
        ("ready", "Lista"),
        ("closed", "Cerrada"),
        ("cancelled", "Cancelada"),
    ]:
        ws2.append([label, by_status.get(st_key, 0)])

    ws2.append(["Total", summary.get("total", 0)])
    ws2.append(["Total estimado (CRC)", summary.get("total_estimated", "0")])

    for col in ws2.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"reporte_ordenes_{date_from}_al_{date_to}.xlsx"
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
